"""Build a single-sheet Excel workbook of per-switch VLAN matrices."""

from __future__ import annotations

import re
from datetime import datetime, timezone
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from swi_mgmt.config import SwitchConfig
from swi_mgmt.models.switch import PortStatus, SwitchSnapshot

# Compact equal width for every VLAN assignment column.
_VLAN_COL_WIDTH = 3.6
_PORT_COL_WIDTH = 4.2
_COMMENT_COL_WIDTH = 18


def _short_port_label(port: PortStatus) -> int | str:
    """Prefer the front-panel / ifIndex number so the cell stays narrow."""
    name = (port.name or "").strip()
    patterns = (
        r"^(\d+)$",
        r"^\d+/(\d+)$",
        r"^port\s+(\d+)\b",
        r"^port[\s_-]*(\d+)\b",
        r"(?:gigabit|ten-?gigabit|fast)?\s*ethernet\s+(?:\d+/)*(\d+)\s*$",
        r"^(?:gi|te|fa|sfp\+?)\s*(?:\d+/)*(\d+)\s*$",
        r"^(?:gigabit|ten-?gigabit|fast)?ethernet\s*\d+(?:/\d+)*/(\d+)$",
        r"(?:copper|fiber|fibre)\s+(?:ethernet\s+)?(?:\d+/)*(\d+)\s*$",
        r"/(\d+)\s*$",
    )
    for pat in patterns:
        m = re.search(pat, name, flags=re.IGNORECASE)
        if m:
            return int(m.group(1))
    return port.index


def _port_untagged(port: PortStatus) -> list[int]:
    if port.untagged_vlans:
        return list(port.untagged_vlans)
    if port.primary_vlan in port.tagged_vlans:
        return []
    return [port.primary_vlan] if port.primary_vlan else []


def _is_trunk(port: PortStatus) -> bool:
    untagged = _port_untagged(port)
    tagged = [v for v in port.tagged_vlans if v not in untagged]
    return bool(tagged) or len(set(untagged) | set(port.tagged_vlans)) > 1


def _hex_fill(color: str) -> PatternFill | None:
    c = (color or "").strip().lstrip("#")
    if len(c) != 6 or any(ch not in "0123456789abcdefABCDEF" for ch in c):
        return None
    return PatternFill(fill_type="solid", fgColor=c.upper())


def _port_config_summary(snap: SwitchSnapshot) -> str:
    ports = snap.ports
    total = len(ports)
    up = sum(1 for p in ports if p.oper_status.name == "UP")
    admin_down = sum(1 for p in ports if p.admin_status.name == "DOWN")
    trunks = sum(1 for p in ports if _is_trunk(p))
    access = total - trunks
    parts = [f"{total} ports", f"{up} up", f"{trunks} trunk", f"{access} access"]
    if admin_down:
        parts.insert(2, f"{admin_down} admin-down")
    return " · ".join(parts)


def _switch_configured_vlans(snap: SwitchSnapshot) -> set[int]:
    return {v.vlan_id for v in snap.vlans}


def _device_name(cfg: SwitchConfig, snap: SwitchSnapshot | None) -> str:
    return (cfg.name or (snap.identity.sys_name if snap else "") or cfg.host).strip()


def _device_details(cfg: SwitchConfig, snap: SwitchSnapshot | None, error: str | None) -> str:
    url = f"http://{cfg.host}/"
    if error:
        return f"{url} · {cfg.host} · {error}"
    assert snap is not None
    id_ = snap.identity
    type_label = " ".join(
        p for p in (id_.vendor, id_.model or id_.sys_descr) if p
    ).strip() or (id_.driver_id or "unknown")
    return f"{url} · {cfg.host} · {type_label} · {_port_config_summary(snap)}"


def _write_vlan_headers(
    ws,
    *,
    id_row: int,
    name_row: int,
    session_vlans: list[dict],
    styles: dict,
) -> int:
    """Write the shared Comment/VLAN header (port cols unlabeled). Returns trailing port column."""
    thin = styles["thin"]
    head_font = styles["head_font"]
    id_font = styles["id_font"]
    name_font = styles["name_font"]
    center = styles["center"]
    name_align = styles["name_align"]
    grey = styles["grey"]

    n_vlans = len(session_vlans)
    trailing_port_col = 3 + n_vlans

    # Port columns stay unlabeled so they can stay narrow; Comment keeps a label.
    for col, text in ((1, ""), (2, "Comment"), (trailing_port_col, "")):
        ws.merge_cells(
            start_row=id_row, start_column=col, end_row=name_row, end_column=col
        )
        cell = ws.cell(row=id_row, column=col, value=text or None)
        cell.font = head_font
        cell.alignment = center
        cell.border = thin
        cell.fill = grey
        ws.cell(row=name_row, column=col).border = thin
        ws.cell(row=name_row, column=col).fill = grey

    for col_i, vlan in enumerate(session_vlans, start=3):
        vid = int(vlan["vlan_id"])
        name = (vlan.get("name") or "").strip()
        fill = _hex_fill(str(vlan.get("color") or "")) or grey

        id_cell = ws.cell(row=id_row, column=col_i, value=vid)
        id_cell.font = id_font
        id_cell.alignment = center
        id_cell.border = thin
        id_cell.fill = fill

        name_cell = ws.cell(row=name_row, column=col_i, value=name)
        name_cell.font = name_font
        name_cell.alignment = name_align
        name_cell.border = thin
        name_cell.fill = fill

    ws.row_dimensions[id_row].height = 18
    ws.row_dimensions[name_row].height = 90
    return trailing_port_col


def _write_switch_section(
    ws,
    *,
    start_row: int,
    cfg: SwitchConfig,
    snap: SwitchSnapshot | None,
    error: str | None,
    session_vlans: list[dict],
    trailing_port_col: int,
    styles: dict,
) -> int:
    """Two-line heading + port rows. Returns next free row."""
    thin = styles["thin"]
    left = styles["left"]
    center = styles["center"]

    name_row = start_row
    detail_row = start_row + 1

    name_cell = ws.cell(row=name_row, column=1, value=_device_name(cfg, snap))
    name_cell.font = Font(bold=True, size=12, color="101828")
    name_cell.alignment = left

    detail = _device_details(cfg, snap, error)
    detail_cell = ws.cell(row=detail_row, column=1, value=detail)
    detail_cell.font = Font(color="475467", size=10)
    detail_cell.alignment = left
    if not error:
        # Make the URL portion clickable via a hyperlink on the detail cell.
        detail_cell.hyperlink = f"http://{cfg.host}/"

    if error or snap is None:
        return detail_row + 2

    configured = _switch_configured_vlans(snap)
    data_start = detail_row + 1

    for r_i, port in enumerate(sorted(snap.ports, key=lambda p: p.index)):
        excel_row = data_start + r_i
        untagged = set(_port_untagged(port))
        tagged = set(port.tagged_vlans) - untagged

        label = _short_port_label(port)
        port_cell = ws.cell(row=excel_row, column=1, value=label)
        port_cell.font = Font(bold=True)
        port_cell.alignment = center
        port_cell.border = thin

        comment_cell = ws.cell(row=excel_row, column=2, value="")
        comment_cell.border = thin
        comment_cell.alignment = left

        for col_i, vlan in enumerate(session_vlans, start=3):
            vid = int(vlan["vlan_id"])
            cell = ws.cell(row=excel_row, column=col_i, value="")
            cell.alignment = center
            cell.border = thin
            if vid not in configured:
                continue
            if vid in untagged:
                cell.value = "(u)"
                fill = _hex_fill(str(vlan.get("color") or ""))
                if fill is not None:
                    cell.fill = fill
            elif vid in tagged:
                cell.value = "t"
                fill = _hex_fill(str(vlan.get("color") or ""))
                if fill is not None:
                    cell.fill = fill

        trail = ws.cell(row=excel_row, column=trailing_port_col, value=label)
        trail.font = Font(bold=True)
        trail.alignment = center
        trail.border = thin

    return data_start + max(1, len(snap.ports)) + 1  # one blank before next switch


def build_vlan_matrix_workbook(
    *,
    switches: Iterable[SwitchConfig],
    snapshots: dict[str, SwitchSnapshot],
    session_vlans: list[dict],
    errors: dict[str, str] | None = None,
) -> bytes:
    """
    One sheet: shared VLAN headings once at the top, then each switch with a
    two-line heading (name, then URL/IP/type/port-config) and port rows.
    Cells are ``(u)`` / ``t``; VLANs not configured on a switch leave that
    column blank. Frozen panes keep Port/Comment + VLAN headers visible.
    """
    errors = errors or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "VLAN Matrix"

    thin = Border(
        left=Side(style="thin", color="D0D5DD"),
        right=Side(style="thin", color="D0D5DD"),
        top=Side(style="thin", color="D0D5DD"),
        bottom=Side(style="thin", color="D0D5DD"),
    )
    styles = {
        "thin": thin,
        "head_font": Font(bold=True, color="0F172A", size=11),
        "id_font": Font(bold=True, color="0F172A", size=11),
        "name_font": Font(color="0F172A", size=9),
        "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "left": Alignment(horizontal="left", vertical="center"),
        "name_align": Alignment(
            horizontal="center", vertical="bottom", textRotation=90, wrap_text=True
        ),
        "grey": PatternFill("solid", fgColor="E4E7EC"),
    }

    ws["A1"] = "SWI-MGMT VLAN matrix export"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (
        f"Exported {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')} UTC · "
        "(u) = untagged / native · t = tagged · blank = not on port or VLAN not on switch"
    )
    ws["A2"].font = Font(color="475467", size=10)

    id_row, name_row = 4, 5
    trailing_port_col = _write_vlan_headers(
        ws,
        id_row=id_row,
        name_row=name_row,
        session_vlans=session_vlans,
        styles=styles,
    )

    row = 6
    switch_list = list(switches)
    if not switch_list:
        ws.cell(row=row, column=1, value="No switches configured.")
    else:
        for cfg in switch_list:
            snap = snapshots.get(cfg.host)
            row = _write_switch_section(
                ws,
                start_row=row,
                cfg=cfg,
                snap=snap,
                error=None
                if snap
                else errors.get(cfg.host, "No snapshot loaded for this switch."),
                session_vlans=session_vlans,
                trailing_port_col=trailing_port_col,
                styles=styles,
            )

    # Freeze Port + Comment columns and the shared VLAN heading rows.
    ws.freeze_panes = "C6"

    ws.column_dimensions["A"].width = _PORT_COL_WIDTH
    ws.column_dimensions["B"].width = _COMMENT_COL_WIDTH
    n_vlans = len(session_vlans)
    for col_i in range(3, 3 + n_vlans):
        ws.column_dimensions[get_column_letter(col_i)].width = _VLAN_COL_WIDTH
    if n_vlans:
        ws.column_dimensions[get_column_letter(trailing_port_col)].width = _PORT_COL_WIDTH

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
